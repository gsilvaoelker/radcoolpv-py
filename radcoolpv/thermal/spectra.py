"""Environmental spectra: atmospheric transmission and the AM1.5 solar spectrum.

Ports ``atmosphericData.m`` and ``solarSpectrum.m``. Both load tabulated data
and interpolate it onto the simulation wavelength grid, setting out-of-range
samples to zero (the MATLAB ``interp1(...); x(isnan)=0`` convention).
"""

from __future__ import annotations

from dataclasses import dataclass

import numpy as np

from .._compat import trapz
from ..constants import CLIGHT, HPLANCK, MICRONTONANO


def load_atmosphere(path: str, lambda_um: np.ndarray) -> np.ndarray:
    """Atmospheric transmittance interpolated onto ``lambda_um`` (0 out of range).

    Port of ``atmosphericData.m`` (Cerro Pachon sky transmission, 7-26 um).
    Columns of the .dat file are ``lambda_um, transmittance``.
    """
    data = np.loadtxt(path)
    lam, tau = data[:, 0], data[:, 1]
    atm = np.interp(lambda_um, lam, tau, left=np.nan, right=np.nan)
    return np.nan_to_num(atm, nan=0.0)


@dataclass
class SolarSpectrum:
    lambda_um: np.ndarray            # simulation grid
    irradiance_per_um: np.ndarray    # AM1.5G, W/(m^2 um), interpolated, 0 out of range
    photon_flux: np.ndarray          # photons/(s m^3), interpolated, 0 out of range
    total_am15: float                # int of irradiance, W/m^2 (~1000)
    # raw (un-interpolated) curve, for plotting:
    raw_lambda_um: np.ndarray
    raw_irradiance_per_um: np.ndarray


def load_solar(path: str, lambda_um: np.ndarray) -> SolarSpectrum:
    """AM1.5G solar spectrum from ``astmg173.xlsx`` (sheet SMARTS2).

    Port of ``solarSpectrum.m``. Column 1 is wavelength in nm, column 3 is the
    global tilt irradiance in W/(m^2 nm).
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["SMARTS2"]
    rows = []
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            continue
        try:
            lam_nm = float(row[0])
            irr = float(row[2])
        except (TypeError, ValueError):
            continue
        rows.append((lam_nm, irr))
    wb.close()
    arr = np.array(rows)

    # arr columns: [xlsx col 1 = wavelength nm, xlsx col 3 = global-tilt irradiance W/m2/nm].
    raw_lambda_um = arr[:, 0] / MICRONTONANO             # nm -> um
    raw_irr_per_um = arr[:, 1] * MICRONTONANO            # W/m2/nm -> W/m2/um

    irr_interp = np.interp(lambda_um, raw_lambda_um, raw_irr_per_um, left=np.nan, right=np.nan)
    irr_interp = np.nan_to_num(irr_interp, nan=0.0)

    # Photon flux (1/(s m^3)) = irradiance * lambda / (h c); interpolate then zero NaN.
    photon_flux_raw = raw_irr_per_um * raw_lambda_um / (HPLANCK * CLIGHT)
    photon_flux = np.interp(lambda_um, raw_lambda_um, photon_flux_raw, left=np.nan, right=np.nan)
    photon_flux = np.nan_to_num(photon_flux, nan=0.0)

    total_am15 = float(trapz(irr_interp, lambda_um))

    return SolarSpectrum(
        lambda_um=lambda_um, irradiance_per_um=irr_interp, photon_flux=photon_flux,
        total_am15=total_am15, raw_lambda_um=raw_lambda_um, raw_irradiance_per_um=raw_irr_per_um,
    )
